from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
from django.db.models.functions import Coalesce
from decimal import Decimal
from datetime import date

from software.models.VentaModel import Venta
from software.models.comprasModel import Compras
from software.models.detalletipousuarioxmodulosModel import Detalletipousuarioxmodulos

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm


def _get_registros(fecha_inicio, fecha_fin, operacion):
    ventas = []
    compras = []

    if operacion in ('todos', 'ventas'):
        qs = Venta.objects.filter(estado=1).select_related(
            'idnumserie__idtipodocumento'
        )
        if fecha_inicio:
            qs = qs.filter(fechaemision__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fechaemision__lte=fecha_fin)
        for v in qs:
            ventas.append({
                'fecha': v.fechaemision,
                'tipo_doc': v.idnumserie.idtipodocumento.nombredocumento,
                'serie_numero': f"{v.idnumserie.numserie}-{v.numcorrelativo}",
                'operacion': 'Venta',
                'entrada': Decimal('0'),
                'salida': v.total_a_pagar,
            })

    if operacion in ('todos', 'compras'):
        qs = Compras.objects.filter(estado=1).annotate(
            total=Sum('compradetalle__subtotal')
        )
        if fecha_inicio:
            qs = qs.filter(fechacompra__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fechacompra__lte=fecha_fin)
        for c in qs:
            compras.append({
                'fecha': c.fechacompra,
                'tipo_doc': 'Comprobante de compra',
                'serie_numero': c.numcorrelativo,
                'operacion': 'Compra',
                'entrada': Decimal(str(c.total or 0)),
                'salida': Decimal('0'),
            })

    # Compras (0) antes que Ventas (1) en el mismo día para el saldo acumulado
    registros = sorted(ventas + compras, key=lambda x: (x['fecha'], 0 if x['operacion'] == 'Compra' else 1))

    saldo = Decimal('0')
    for r in registros:
        saldo += r['entrada'] - r['salida']
        r['saldo_final'] = saldo

    return registros


def registro(request):
    id2 = request.session.get('idtipousuario')
    if not id2:
        return render(request, 'errors/error.html')

    permisos = Detalletipousuarioxmodulos.objects.filter(idtipousuario=id2)

    fecha_inicio = request.GET.get('fecha_inicio') or None
    fecha_fin = request.GET.get('fecha_fin') or None
    operacion = request.GET.get('operacion', 'todos')

    registros = _get_registros(fecha_inicio, fecha_fin, operacion)

    total_entradas = sum(r['entrada'] for r in registros)
    total_salidas = sum(r['salida'] for r in registros)
    saldo_total = total_entradas - total_salidas

    return render(request, 'registro/registro.html', {
        'permisos': permisos,
        'registros': registros,
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'operacion': operacion,
        'total_entradas': total_entradas,
        'total_salidas': total_salidas,
        'saldo_total': saldo_total,
        'nombrecompleto': request.session.get('nombrecompleto'),
    })


def export_registro_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio') or None
    fecha_fin = request.GET.get('fecha_fin') or None
    operacion = request.GET.get('operacion', 'todos')

    registros = _get_registros(fecha_inicio, fecha_fin, operacion)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["#", "Fecha", "Tipo Doc.", "Serie/Número", "Operación",
               "Entradas (S/)", "Salidas (S/)", "Saldo Final (S/)"]
    ws.append(headers)

    for i, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    col_widths = [5, 14, 25, 18, 12, 16, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, r in enumerate(registros, 1):
        ws.append([
            idx,
            r['fecha'].strftime('%d/%m/%Y'),
            r['tipo_doc'],
            r['serie_numero'],
            r['operacion'],
            float(r['entrada']) if r['entrada'] else '-',
            float(r['salida']) if r['salida'] else '-',
            float(r['saldo_final']),
        ])
        row = ws[ws.max_row]
        if r['operacion'] == 'Compra':
            row[5].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        else:
            row[6].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    total_row = ws.max_row + 1
    ws.cell(total_row, 1, "TOTALES").font = Font(bold=True)
    ws.cell(total_row, 6, float(sum(r['entrada'] for r in registros))).font = Font(bold=True)
    ws.cell(total_row, 7, float(sum(r['salida'] for r in registros))).font = Font(bold=True)
    ws.cell(total_row, 8, float(sum(r['entrada'] - r['salida'] for r in registros))).font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=registro.xlsx'
    wb.save(response)
    return response


def export_registro_pdf(request):
    fecha_inicio = request.GET.get('fecha_inicio') or None
    fecha_fin = request.GET.get('fecha_fin') or None
    operacion = request.GET.get('operacion', 'todos')

    registros = _get_registros(fecha_inicio, fecha_fin, operacion)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=registro.pdf'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph("<b>Registro de Ventas y Compras</b>", styles['Title'])
    elements.append(title)
    if fecha_inicio or fecha_fin:
        sub = Paragraph(
            f"Período: {fecha_inicio or '—'} a {fecha_fin or '—'}", styles['Normal']
        )
        elements.append(sub)
    elements.append(Spacer(1, 0.5*cm))

    data = [["#", "Fecha", "Tipo Doc.", "Serie/Número", "Operación",
             "Entradas (S/)", "Salidas (S/)", "Saldo Final (S/)"]]

    for idx, r in enumerate(registros, 1):
        data.append([
            str(idx),
            r['fecha'].strftime('%d/%m/%Y'),
            r['tipo_doc'],
            r['serie_numero'],
            r['operacion'],
            f"{r['entrada']:,.2f}" if r['entrada'] else '-',
            f"{r['salida']:,.2f}" if r['salida'] else '-',
            f"{r['saldo_final']:,.2f}",
        ])

    total_ent = sum(r['entrada'] for r in registros)
    total_sal = sum(r['salida'] for r in registros)
    data.append(["", "", "", "", "TOTAL",
                 f"{total_ent:,.2f}", f"{total_sal:,.2f}",
                 f"{total_ent - total_sal:,.2f}"])

    col_widths = [1*cm, 2.5*cm, 5*cm, 3.5*cm, 2.5*cm, 3*cm, 3*cm, 3*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F2F2')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWHEIGHT', (0, 0), (-1, -1), 18),
    ]))

    elements.append(table)
    doc.build(elements)
    return response
