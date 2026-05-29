from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal, ROUND_HALF_UP

from software.models.VentaModel import Venta
from software.models.comprasModel import Compras
from software.models.detalletipousuarioxmodulosModel import Detalletipousuarioxmodulos

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

ZERO = Decimal('0')
CENT = Decimal('0.01')


def _q(val):
    return val.quantize(CENT, ROUND_HALF_UP)


def _get_registros_val(fecha_inicio, fecha_fin, operacion):
    ventas = []
    compras = []

    if operacion in ('todos', 'ventas'):
        qs = Venta.objects.filter(estado=1).select_related(
            'idnumserie__idtipodocumento'
        ).annotate(
            total_cant=Coalesce(Sum('ventadetalle__cantidad'), Value(0), output_field=DecimalField()),
            total_monto=Coalesce(Sum('ventadetalle__preciosubtotal'), Value(0), output_field=DecimalField()),
        )
        if fecha_inicio:
            qs = qs.filter(fechaemision__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fechaemision__lte=fecha_fin)
        for v in qs:
            cant  = v.total_cant or ZERO
            total = v.total_monto or ZERO
            unit  = _q(total / cant) if cant else ZERO
            ventas.append({
                'fecha': v.fechaemision,
                'tipo_doc': v.idnumserie.idtipodocumento.nombredocumento,
                'serie_numero': f"{v.idnumserie.numserie}-{v.numcorrelativo}",
                'operacion': 'Venta',
                'ent_cant': None, 'ent_unit': None, 'ent_total': None,
                'sal_cant': cant,  'sal_unit': unit,  'sal_total': total,
            })

    if operacion in ('todos', 'compras'):
        qs = Compras.objects.filter(estado=1).annotate(
            total_cant=Sum('compradetalle__cantidad'),
            total_monto=Sum('compradetalle__subtotal'),
        )
        if fecha_inicio:
            qs = qs.filter(fechacompra__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fechacompra__lte=fecha_fin)
        for c in qs:
            cant  = Decimal(str(c.total_cant  or 0))
            total = Decimal(str(c.total_monto or 0))
            unit  = _q(total / cant) if cant else ZERO
            compras.append({
                'fecha': c.fechacompra,
                'tipo_doc': 'Comprobante',
                'serie_numero': c.numcorrelativo,
                'operacion': 'Compra',
                'ent_cant': cant, 'ent_unit': unit, 'ent_total': total,
                'sal_cant': None, 'sal_unit': None, 'sal_total': None,
            })

    registros = sorted(
        ventas + compras,
        key=lambda x: (x['fecha'], 0 if x['operacion'] == 'Compra' else 1)
    )

    # Calcular saldo acumulado (método costo promedio ponderado)
    saldo_cant  = ZERO
    saldo_total = ZERO

    for r in registros:
        if r['operacion'] == 'Compra':
            saldo_cant  += r['ent_cant']
            saldo_total += r['ent_total']
        else:
            saldo_unit_prev = _q(saldo_total / saldo_cant) if saldo_cant else ZERO
            costo_salida    = _q(r['sal_cant'] * saldo_unit_prev)
            r['sal_unit']   = saldo_unit_prev   # valor al costo, no al precio de venta
            r['sal_total']  = costo_salida
            saldo_cant  -= r['sal_cant']
            saldo_total -= costo_salida

        saldo_unit = _q(saldo_total / saldo_cant) if saldo_cant else ZERO
        r['saldo_cant']  = saldo_cant
        r['saldo_unit']  = saldo_unit
        r['saldo_total'] = saldo_total

    return registros


def registro_valorizado(request):
    id2 = request.session.get('idtipousuario')
    if not id2:
        return render(request, 'errors/error.html')

    permisos     = Detalletipousuarioxmodulos.objects.filter(idtipousuario=id2)
    fecha_inicio = request.GET.get('fecha_inicio') or None
    fecha_fin    = request.GET.get('fecha_fin')    or None
    operacion    = request.GET.get('operacion', 'todos')

    registros = _get_registros_val(fecha_inicio, fecha_fin, operacion)

    return render(request, 'registroValorizado/registroValorizado.html', {
        'permisos':       permisos,
        'registros':      registros,
        'fecha_inicio':   fecha_inicio or '',
        'fecha_fin':      fecha_fin    or '',
        'operacion':      operacion,
        'nombrecompleto': request.session.get('nombrecompleto'),
    })


def export_valorizado_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio') or None
    fecha_fin    = request.GET.get('fecha_fin')    or None
    operacion    = request.GET.get('operacion', 'todos')
    registros    = _get_registros_val(fecha_inicio, fecha_fin, operacion)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro Valorizado"

    h_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    g_fill   = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    r_fill   = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    b_fill   = PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid")
    h_font   = Font(color="FFFFFF", bold=True)
    bold     = Font(bold=True)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fila 1: grupos
    ws.merge_cells('A1:D1'); ws['A1'] = ''
    ws.merge_cells('E1:G1'); ws['E1'] = 'ENTRADA'
    ws.merge_cells('H1:J1'); ws['H1'] = 'SALIDA'
    ws.merge_cells('K1:M1'); ws['K1'] = 'SALDO'
    for col, fill in [('E', g_fill), ('H', r_fill), ('K', b_fill)]:
        ws[col + '1'].fill = fill
        ws[col + '1'].font = bold
        ws[col + '1'].alignment = center

    # Fila 2: subencabezados
    headers = ['#', 'Fecha', 'Tipo Doc.', 'Serie/Número',
               'Cantidad', 'Valor (S/)', 'Total (S/)',
               'Cantidad', 'Valor (S/)', 'Total (S/)',
               'Cantidad', 'Valor (S/)', 'Total (S/)']
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = center

    widths = [4, 13, 20, 16, 10, 12, 12, 10, 12, 12, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def val(v): return float(v) if v is not None else '-'

    for idx, r in enumerate(registros, 1):
        ws.append([
            idx,
            r['fecha'].strftime('%d/%m/%Y'),
            r['tipo_doc'], r['serie_numero'],
            val(r['ent_cant']),  val(r['ent_unit']),  val(r['ent_total']),
            val(r['sal_cant']),  val(r['sal_unit']),  val(r['sal_total']),
            float(r['saldo_cant']), float(r['saldo_unit']), float(r['saldo_total']),
        ])
        row = ws[ws.max_row]
        for c in range(4, 7):   row[c].fill = g_fill  # Entrada  (cols E-G, idx 4-6)
        for c in range(7, 10):  row[c].fill = r_fill  # Salida   (cols H-J, idx 7-9)
        for c in range(10, 13): row[c].fill = b_fill  # Saldo    (cols K-M, idx 10-12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=registro_valorizado.xlsx'
    wb.save(response)
    return response


def export_valorizado_pdf(request):
    fecha_inicio = request.GET.get('fecha_inicio') or None
    fecha_fin    = request.GET.get('fecha_fin')    or None
    operacion    = request.GET.get('operacion', 'todos')
    registros    = _get_registros_val(fecha_inicio, fecha_fin, operacion)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=registro_valorizado.pdf'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            leftMargin=0.8*cm, rightMargin=0.8*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm)
    styles   = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("<b>Registro Valorizado</b>", styles['Title']))
    if fecha_inicio or fecha_fin:
        elements.append(Paragraph(
            f"Período: {fecha_inicio or '—'} al {fecha_fin or '—'}", styles['Normal']))
    elements.append(Spacer(1, 0.4*cm))

    GREEN = colors.HexColor('#D1E7DD')
    RED   = colors.HexColor('#F8D7DA')
    BLUE  = colors.HexColor('#CFE2FF')
    NAVY  = colors.HexColor('#1F4E79')

    def fv(v): return f"{v:,.2f}" if v is not None else '-'

    # Encabezado doble
    data = [
        ['', '', '', '', 'ENTRADA', '', '', 'SALIDA', '', '', 'SALDO', '', ''],
        ['#', 'Fecha', 'Tipo Doc.', 'Serie/Nro',
         'Cant.', 'Valor\nS/', 'Total\nS/',
         'Cant.', 'Valor\nS/', 'Total\nS/',
         'Cant.', 'Valor\nS/', 'Total\nS/'],
    ]

    for idx, r in enumerate(registros, 1):
        data.append([
            str(idx),
            r['fecha'].strftime('%d/%m/%Y'),
            r['tipo_doc'], r['serie_numero'],
            fv(r['ent_cant']),  fv(r['ent_unit']),  fv(r['ent_total']),
            fv(r['sal_cant']),  fv(r['sal_unit']),  fv(r['sal_total']),
            fv(r['saldo_cant']), fv(r['saldo_unit']), fv(r['saldo_total']),
        ])

    cw = [0.7*cm, 2*cm, 3.5*cm, 2.8*cm,
          1.8*cm, 2*cm, 2*cm,
          1.8*cm, 2*cm, 2*cm,
          1.8*cm, 2*cm, 2*cm]

    table = Table(data, colWidths=cw, repeatRows=2)
    table.setStyle(TableStyle([
        # Grupo headers (fila 0)
        ('SPAN',        (4, 0), (6, 0)),  # ENTRADA
        ('SPAN',        (7, 0), (9, 0)),  # SALIDA
        ('SPAN',        (10, 0),(12, 0)), # SALDO
        ('BACKGROUND',  (4, 0), (6, 0),  GREEN),
        ('BACKGROUND',  (7, 0), (9, 0),  RED),
        ('BACKGROUND',  (10,0), (12, 0), BLUE),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        # Fila subencabezados (fila 1)
        ('BACKGROUND',  (0, 1), (-1, 1),  NAVY),
        ('TEXTCOLOR',   (0, 1), (-1, 1),  colors.white),
        ('FONTNAME',    (0, 1), (-1, 1),  'Helvetica-Bold'),
        # General
        ('FONTSIZE',    (0, 0), (-1, -1), 7),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWHEIGHT',   (0, 0), (-1, -1), 16),
    ]))
    elements.append(table)
    doc.build(elements)
    return response
